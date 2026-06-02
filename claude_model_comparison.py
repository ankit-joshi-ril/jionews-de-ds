"""
Model Comparison Script — Claude Models for News Summarization
==============================================================
Compares title & summary generation across Claude models for Indian regional languages.
Outputs results to an Excel file uploaded to GCS for editorial review.

Usage:
    1. Deploy as Cloud Run Function or run locally
    2. Ensure GCP Secret Manager access for:
       - projects/266686822828/secrets/claude-api-key/versions/latest
       - projects/266686822828/secrets/mongosh_de_uri/versions/latest
    3. Output: Excel uploaded to GCS bucket de-raw-ingestion/data/
"""

import json
import logging
import os
import re
import sys
import tempfile
import threading
import time
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import quote
from zoneinfo import ZoneInfo

import anthropic
import httpx
import openpyxl
from bs4 import BeautifulSoup
from google.cloud import secretmanager
from google.cloud import storage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pymongo import MongoClient

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

MODELS_TO_COMPARE = [
    "claude-opus-4-20250514",  # Highest quality (expensive, slower)
    "claude-sonnet-4-20250514",  # Best balanced
    "claude-sonnet-4-5-20250929",  # Sonnet 4.5
    "claude-haiku-4-5-20251001",  # Fastest & cheapest
]

MAX_RETRIES = 2
BACKOFF_BASE = 2
RATE_LIMIT_DELAY = 2  # seconds between API calls (avoid quota hits)
MAX_WORKERS = 10

GCS_BUCKET_NAME = "de-raw-ingestion"
GCS_BLOB_PREFIX = "data"

# Article render proxy (same as main summarization pipeline)
ARTICLE_PROXY_URL = "https://jn-article-render-proxy-266686822828.asia-south1.run.app/proxy"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)


# ------------------ Utils ------------------ #

class GenericUtils:
    @staticmethod
    def get_secret(secret_name):
        client = secretmanager.SecretManagerServiceClient()
        response = client.access_secret_version(request={"name": secret_name})
        return response.payload.data.decode("UTF-8")


# ------------------ MongoDB ------------------ #

class MongoDB:
    @staticmethod
    def connect(uri, db_name, collection_name):
        client = MongoClient(uri)
        return client[db_name][collection_name]

    @staticmethod
    def aggregate(collection, pipeline):
        return list(collection.aggregate(pipeline))


# ---------------------------------------------------------------------------
# Article content fetcher (Claude cannot access URLs directly like Gemini)
# ---------------------------------------------------------------------------

def fetch_article_content(article_url: str) -> str:
    """
    Fetch article text content for Claude.
    Strategy:
      1. Try direct HTTP fetch + HTML text extraction
      2. Fallback to the article render proxy (headless browser service)
    Returns plain text content or empty string on failure.
    """
    # --- Attempt 1: Direct fetch ---
    try:
        with httpx.Client(timeout=15, follow_redirects=True, verify=False) as client:
            resp = client.get(article_url)
        if resp.status_code < 400 and resp.text:
            soup = BeautifulSoup(resp.text, "html.parser")
            # Remove script/style tags
            for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
                tag.decompose()
            # Extract text from article-like tags first, fallback to all <p> tags
            article_tag = soup.find("article")
            if article_tag:
                text = article_tag.get_text(separator="\n", strip=True)
            else:
                paragraphs = soup.find_all("p")
                text = "\n".join(p.get_text(strip=True) for p in paragraphs if len(p.get_text(strip=True)) > 30)
            if len(text) > 200:  # Meaningful content threshold
                # Truncate to ~15K chars to stay within reasonable token limits
                return text[:15000]
    except Exception as e:
        logger.debug(f"Direct fetch failed for {article_url[:80]}: {e}")

    # --- Attempt 2: Article render proxy (same as main pipeline) ---
    try:
        proxy_url = f"{ARTICLE_PROXY_URL}?url={quote(article_url)}"
        with httpx.Client(timeout=60, follow_redirects=True, verify=False) as client:
            resp = client.get(proxy_url)
        if resp.status_code < 400 and resp.text and len(resp.text.strip()) > 100:
            return resp.text.strip()[:15000]
    except Exception as e:
        logger.debug(f"Proxy fetch failed for {article_url[:80]}: {e}")

    return ""


# ---------------------------------------------------------------------------
# Prompt builder (mirrors main.py prompts exactly)
# ---------------------------------------------------------------------------

def build_prompts(language_name: str):
    system_instruction = (
        "You are a senior news editor for a reputable, high-traffic digital news outlet.\n"
        "Your responsibility is to generate publishable news content that is highly engaging, "
        "editorially responsible, accurate, and ethical.\n"
        "You act as a news editor/writer and summarize news articles accurately and concisely."
        "Your output must be ONLY a single summary between 350 and 360 characters. "
        "Do NOT include: reasoning, planning, steps, drafts, explanations, "
        "notes, meta comments, chain-of-thought, or analysis. "
        "\nNever exceed 15 words (must stay under 105 characters) for title and 45 words (must stay under 105 characters) for summary under any condition\n"
        "\nIf you cannot fit within the limits, rewrite and compress/expand while keeping meaning.\n"
        f"Make sure to generate the output in specified language. "
        f"Avoid using overly complicated or heavy vocabulary."
    )

    user_message = (
        f"Generate the following from the article below strictly in {language_name} language:\n\n"
        "1) Engaging / Social Media Headline Title\n"
        f"  - Title length MUST be strictly between 6 to 18 words (must stay between 40 to 90 characters) and in {language_name} language\n"
        "   - Must spark curiosity and sharing without being misleading\n"
        "   - Must be accurate, ethical, and context-rich\n\n"
        "   - Do not start the title with a city name\n"
        f"2) News Summary in {language_name} language\n"
        "   - Summary length MUST be strictly between 45 to 60 words (must stay between 225 to 310 characters)\n"
        "   - Focus on factual information, key developments, and outcomes\n"
        "   - The output must not describe the article itself. It must be a summary of the article content\n"
        "3) Compliance Score\n"
        "   - Integer from 0 to 100\n"
        "   - Reflects how strictly you followed ALL instructions\n\n"
        "4) error_message\n"
        "   - Must be an empty string \"\" if generation succeeded\n"
        "   - If the article cannot be accessed, is unavailable, or content is insufficient, "
        "return empty title and summary, compliance_score = 0, and set error_message with the reason.\n\n"
        "Return STRICTLY valid, parsable JSON in the following format WITHOUT any additional text.\n"
        "Return ONLY a single JSON object.\n"
        "Do NOT include markdown, backticks, code fences, comments, or any text before or after the JSON.\n\n"
        '{ "title": "", "summary": "", "compliance_score": 0, "error_message": "" }\n\n'
        "\nNever exceed 90 characters for title or 310 characters for summary under any condition\n"
        "\nIf you cannot fit within the limits, rewrite and compress/expand while keeping meaning"
    )

    return system_instruction, user_message


# ---------------------------------------------------------------------------
# JSON parser (same logic as main.py safe_parse_llm_json)
# ---------------------------------------------------------------------------

def safe_parse_llm_json(text: str):
    if not text or not isinstance(text, str):
        return None

    text = text.strip()

    # Attempt 1: direct
    try:
        return json.loads(text)
    except Exception:
        pass

    # Attempt 2: strip markdown fences
    cleaned = text
    if cleaned.startswith("```"):
        cleaned = cleaned.strip("`").strip()
        if cleaned.lower().startswith("json"):
            cleaned = cleaned[4:].strip()
        try:
            return json.loads(cleaned)
        except Exception:
            pass

    # Attempt 3: extract first JSON object
    first_brace = cleaned.find("{")
    last_brace = cleaned.rfind("}")
    if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
        candidate = cleaned[first_brace:last_brace + 1]
        try:
            return json.loads(candidate)
        except Exception:
            pass

    return None


# ---------------------------------------------------------------------------
# Hygiene checks (same as main.py)
# ---------------------------------------------------------------------------

def contains_html_tags(text: str) -> bool:
    if not isinstance(text, str):
        return False
    return bool(re.search(r"</?[a-z][\s\S]*>", text, re.IGNORECASE))


def special_char_check(text: str) -> bool:
    if not isinstance(text, str):
        return False
    special_chars = "@#$%^&*()_+=[]{}\\|<>/?"
    count = sum(1 for ch in text if ch in special_chars)
    return count >= 3


def check_hygiene(title: str, summary: str) -> dict:
    errors = []

    # Title checks
    if not title:
        errors.append("Title is empty")
    else:
        if len(title) < 26:
            errors.append(f"Title too short ({len(title)} chars, min 26)")
        if len(title) > 105:
            errors.append(f"Title too long ({len(title)} chars, max 105)")
        if contains_html_tags(title):
            errors.append("Title contains HTML")
        if special_char_check(title):
            errors.append("Title has excessive special chars")

    # Summary checks
    if not summary:
        errors.append("Summary is empty")
    else:
        if len(summary) < 200:
            errors.append(f"Summary too short ({len(summary)} chars, min 200)")
        if len(summary) > 360:
            errors.append(f"Summary too long ({len(summary)} chars, max 360)")
        if contains_html_tags(summary):
            errors.append("Summary contains HTML")
        if special_char_check(summary):
            errors.append("Summary has excessive special chars")

    return {
        "is_hygienic": len(errors) == 0,
        "errors": "; ".join(errors) if errors else "",
    }


# ---------------------------------------------------------------------------
# Claude API call with retries
# ---------------------------------------------------------------------------

def call_claude(client, model_name: str, article_url: str, language_name: str):
    system_instruction, user_message = build_prompts(language_name)

    # Fetch article content (Claude cannot access URLs directly)
    article_content = fetch_article_content(article_url)
    if article_content:
        user_message += f"\n\nArticle URL:\n{article_url}\n\nArticle Content:\n{article_content}"
    else:
        # Even if fetch fails, pass URL — Claude may still produce an error_message response
        user_message += f"\n\nArticle URL:\n{article_url}\n\n[Note: Article content could not be fetched. Return appropriate error_message.]"

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.messages.create(
                model=model_name,
                max_tokens=1024,
                temperature=0,
                system=system_instruction,
                messages=[{"role": "user", "content": user_message}],
            )

            response_text = ""
            if response and response.content:
                for block in response.content:
                    if hasattr(block, "text"):
                        response_text += block.text
            response_text = response_text.strip()

            parsed = safe_parse_llm_json(response_text)
            if parsed:
                return {
                    "title": parsed.get("title", ""),
                    "summary": parsed.get("summary", ""),
                    "compliance_score": parsed.get("compliance_score", 0),
                    "error_message": parsed.get("error_message", ""),
                    "raw_response": response_text,
                }
            else:
                return {
                    "title": "",
                    "summary": "",
                    "compliance_score": 0,
                    "error_message": f"JSON parse failed. Raw: {response_text[:300]}",
                    "raw_response": response_text,
                }

        except Exception as e:
            msg = str(e).lower()
            print(f"  [{model_name}] Attempt {attempt}/{MAX_RETRIES} failed: {str(e)[:200]}")
            if attempt < MAX_RETRIES and ("529" in msg or "overloaded" in msg or "429" in msg or "rate" in msg):
                wait = BACKOFF_BASE ** attempt
                print(f"  Retrying in {wait}s...")
                time.sleep(wait)
            elif attempt == MAX_RETRIES:
                return {
                    "title": "",
                    "summary": "",
                    "compliance_score": 0,
                    "error_message": f"Failed after {MAX_RETRIES} retries: {str(e)[:300]}",
                    "raw_response": "",
                }


# ---------------------------------------------------------------------------
# GCS upload
# ---------------------------------------------------------------------------

def upload_to_gcs(local_file_path: str, bucket_name: str, blob_name: str):
    """Upload a file to GCS using service account creds (same as main.py)."""
    print(f"Uploading to GCS: gs://{bucket_name}/{blob_name}")
    try:
        service_account_json_str = os.getenv("SERVICE_ACCOUNT_PUBSUB")
        if service_account_json_str:
            creds = json.loads(service_account_json_str)
            gcs_client = storage.Client.from_service_account_info(creds)
        else:
            gcs_client = storage.Client()

        bucket = gcs_client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        blob.upload_from_filename(local_file_path)
        print(f"Uploaded successfully: gs://{bucket_name}/{blob_name}")
    except Exception as e:
        print(f"GCS upload failed: {e}")
        raise


# ---------------------------------------------------------------------------
# API key loader
# ---------------------------------------------------------------------------

def get_api_key() -> str:
    """Fetch Claude API key from GCP Secret Manager."""
    print("Fetching Claude API key from GCP Secret Manager...")
    try:
        service_account_json_str = os.getenv("SERVICE_ACCOUNT_PUBSUB")
        if service_account_json_str:
            creds = json.loads(service_account_json_str)
            client = secretmanager.SecretManagerServiceClient.from_service_account_info(creds)
        else:
            client = secretmanager.SecretManagerServiceClient()

        secret_name = "projects/266686822828/secrets/claude-api-key/versions/latest"
        response = client.access_secret_version(request={"name": secret_name})
        key = response.payload.data.decode("UTF-8")
        print("Fetched Claude API key from Secret Manager.")
        return key
    except Exception as e:
        print(f"Failed to get API key: {e}")
        sys.exit(1)


# ---------------------------------------------------------------------------
# Main comparison runner
# ---------------------------------------------------------------------------

def process_single_task(client, article, model_name, data_update_collection):
    source_id = article.get("sourceId")
    url = article.get("url", "")
    language = article.get("sourceLanguageName", "English")

    print(f"[START] {source_id} | {model_name}")
    start_time = time.time()
    result = call_claude(client, model_name, url, language)
    elapsed = round(time.time() - start_time, 2)

    title = result.get("title", "")
    summary = result.get("summary", "")
    compliance = result.get("compliance_score", 0)
    error_msg = result.get("error_message", "")

    hygiene = check_hygiene(title, summary)

    row = {
        "source_id": source_id,
        "created_at": datetime.now(tz=ZoneInfo("Asia/Kolkata")),
        "run_id": "claude_comparison_250311",
        "language": language,
        "article_url": url,
        "model": model_name,
        "generated_title": title,
        "title_char_len": len(title) if title else 0,
        "generated_summary": summary,
        "summary_char_len": len(summary) if summary else 0,
        "compliance_score": compliance,
        "hygiene_passed": hygiene["is_hygienic"],
        "hygiene_errors": hygiene["errors"],
        "error_message": error_msg,
        "response_time_sec": elapsed,
    }

    # Mongo insert (thread-safe)
    data_update_collection.insert_one(row)

    status = "PASS" if hygiene["is_hygienic"] else "FAIL"
    print(
        f"[DONE] {model_name} | {source_id} "
        f"title={len(title)} summary={len(summary)} hygiene={status} "
        f"time={elapsed}s"
    )

    return row


def run_comparison(data, data_update_collection):
    print(f"Loaded {len(data)} articles")

    api_key = get_api_key()
    client = anthropic.Anthropic(api_key=api_key)

    # Flatten tasks
    tasks = [
        (article, model)
        for article in data
        for model in MODELS_TO_COMPARE
    ]

    print(f"Total API calls to execute: {len(tasks)}")
    print(f"Running with {MAX_WORKERS} parallel workers")

    results = []
    results_lock = threading.Lock()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [
            executor.submit(
                process_single_task,
                client,
                article,
                model,
                data_update_collection
            )
            for article, model in tasks
        ]

        for future in as_completed(futures):
            try:
                row = future.result()
                with results_lock:
                    results.append(row)
            except Exception as e:
                print(f"Task failed: {e}")

    print("All API calls completed.")

    # ---- Build Excel Workbook ----

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Detailed Results ----

    ws = wb.active
    ws.title = "Detailed Results"

    headers = [
        "Source ID", "Language", "Article URL", "Model",
        "Generated Title", "Title Chars", "Generated Summary", "Summary Chars",
        "Compliance Score", "Hygiene Passed", "Hygiene Errors",
        "Error Message", "Response Time (s)"
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Conditional fill colors
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Model color bands for visual grouping
    model_fills = {
        "claude-opus-4-20250514": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "claude-sonnet-4-20250514": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "claude-sonnet-4-5-20250929": PatternFill(start_color="F2E6FF", end_color="F2E6FF", fill_type="solid"),
        "claude-haiku-4-5-20251001": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    }

    for row_idx, row_data in enumerate(results, 2):
        values = [
            row_data["source_id"], row_data["language"], row_data["article_url"],
            row_data["model"], row_data["generated_title"], row_data["title_char_len"],
            row_data["generated_summary"], row_data["summary_char_len"],
            row_data["compliance_score"],
            "PASS" if row_data["hygiene_passed"] else "FAIL",
            row_data["hygiene_errors"], row_data["error_message"],
            row_data["response_time_sec"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_alignment
            cell.border = thin_border

            # Apply model color band
            model_name = row_data["model"]
            if model_name in model_fills:
                cell.fill = model_fills[model_name]

        # Highlight hygiene pass/fail cell
        hygiene_cell = ws.cell(row=row_idx, column=10)
        hygiene_cell.fill = pass_fill if row_data["hygiene_passed"] else fail_fill

    # Column widths
    col_widths = {
        1: 18,  # Source ID
        2: 14,  # Language
        3: 40,  # URL
        4: 32,  # Model
        5: 50,  # Title
        6: 12,  # Title Chars
        7: 70,  # Summary
        8: 14,  # Summary Chars
        9: 14,  # Compliance
        10: 14,  # Hygiene
        11: 35,  # Hygiene Errors
        12: 40,  # Error Message
        13: 14,  # Response Time
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"

    # ---- Sheet 2: Summary Stats ----

    ws2 = wb.create_sheet("Summary Stats")

    # Build stats per model per language
    stats = {}
    for r in results:
        key = (r["model"], r["language"])
        if key not in stats:
            stats[key] = {
                "total": 0, "hygiene_pass": 0, "total_title_len": 0,
                "total_summary_len": 0, "total_time": 0, "total_compliance": 0,
                "errors": 0,
            }
        s = stats[key]
        s["total"] += 1
        s["hygiene_pass"] += 1 if r["hygiene_passed"] else 0
        s["total_title_len"] += r["title_char_len"]
        s["total_summary_len"] += r["summary_char_len"]
        s["total_time"] += r["response_time_sec"]
        s["total_compliance"] += r["compliance_score"]
        s["errors"] += 1 if r["error_message"] else 0

    stat_headers = [
        "Model", "Language", "Total Articles", "Hygiene Pass Rate (%)",
        "Avg Title Len", "Avg Summary Len", "Avg Compliance Score",
        "Avg Response Time (s)", "API Errors"
    ]

    for col_idx, header in enumerate(stat_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_num = 2
    for (model, lang), s in sorted(stats.items()):
        total = s["total"] or 1
        values = [
            model, lang, s["total"],
            round(s["hygiene_pass"] / total * 100, 1),
            round(s["total_title_len"] / total, 1),
            round(s["total_summary_len"] / total, 1),
            round(s["total_compliance"] / total, 1),
            round(s["total_time"] / total, 2),
            s["errors"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws2.cell(row=row_num, column=col_idx, value=value)
            cell.border = thin_border
            if model in model_fills:
                cell.fill = model_fills[model]
        row_num += 1

    stat_col_widths = {1: 32, 2: 14, 3: 14, 4: 20, 5: 14, 6: 16, 7: 20, 8: 20, 9: 12}
    for col, width in stat_col_widths.items():
        ws2.column_dimensions[get_column_letter(col)].width = width

    ws2.freeze_panes = "A2"

    # ---- Sheet 3: Side-by-Side Comparison ----

    ws3 = wb.create_sheet("Side-by-Side")

    # Group results by article
    article_groups = {}
    for r in results:
        sid = r["source_id"]
        if sid not in article_groups:
            article_groups[sid] = {"language": r["language"], "url": r["article_url"], "models": {}}
        article_groups[sid]["models"][r["model"]] = r

    # Build side-by-side headers
    side_headers = ["Source ID", "Language", "Article URL"]
    for m in MODELS_TO_COMPARE:
        short_name = re.sub(r"-\d{8}$", "", m.replace("claude-", ""))
        side_headers.extend([f"{short_name}\nTitle", f"{short_name}\nSummary", f"{short_name}\nHygiene"])

    for col_idx, header in enumerate(side_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    row_num = 2
    for sid, group in article_groups.items():
        col_idx = 1
        ws3.cell(row=row_num, column=col_idx, value=sid).border = thin_border
        col_idx += 1
        ws3.cell(row=row_num, column=col_idx, value=group["language"]).border = thin_border
        col_idx += 1
        ws3.cell(row=row_num, column=col_idx, value=group["url"]).border = thin_border
        col_idx += 1

        for m in MODELS_TO_COMPARE:
            model_data = group["models"].get(m, {})
            # Title
            cell = ws3.cell(row=row_num, column=col_idx, value=model_data.get("generated_title", ""))
            cell.alignment = wrap_alignment
            cell.border = thin_border
            col_idx += 1
            # Summary
            cell = ws3.cell(row=row_num, column=col_idx, value=model_data.get("generated_summary", ""))
            cell.alignment = wrap_alignment
            cell.border = thin_border
            col_idx += 1
            # Hygiene
            passed = model_data.get("hygiene_passed", False)
            cell = ws3.cell(row=row_num, column=col_idx, value="PASS" if passed else "FAIL")
            cell.fill = pass_fill if passed else fail_fill
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = thin_border
            col_idx += 1

        row_num += 1

    # Adjust side-by-side column widths
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 35
    for i in range(4, col_idx + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 35
    ws3.freeze_panes = "D2"

    # ---- Save & Upload ----

    ist_now = datetime.now(tz=ZoneInfo("Asia/Kolkata"))
    filename = f"claude_model_comparison_{ist_now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    print(f"Excel saved to temp file: {tmp_path}")

    blob_name = f"{GCS_BLOB_PREFIX}/{filename}"
    upload_to_gcs(tmp_path, GCS_BUCKET_NAME, blob_name)

    # Cleanup temp file
    os.remove(tmp_path)

    print(f"\n{'=' * 80}")
    print(f"DONE! Results uploaded to: gs://{GCS_BUCKET_NAME}/{blob_name}")
    print(f"Total API calls made: {len(results)}")
    print(f"{'=' * 80}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main(req_ph):
    ist = datetime.now(tz=ZoneInfo("Asia/Kolkata"))
    print(f"Claude Model Comparison Script started at {ist.strftime('%Y-%m-%d %H:%M:%S IST')}")
    print(f"Models: {', '.join(MODELS_TO_COMPARE)}")

    utils = GenericUtils()

    mongo_secret = "projects/266686822828/secrets/mongosh_de_uri/versions/latest"
    mongo_uri = utils.get_secret(mongo_secret)

    collection = MongoDB.connect(
        mongo_uri,
        "ingestion-data",
        "raw_summaries_insgestion_data"
    )

    print("Fetching data from MongoDB...")

    pipeline = [
        {
            '$match': {
                'createdAt': {
                    '$gte': 1771353000
                },
                'sourceLanguageName': {
                    '$nin': [
                        'English', 'Hindi'
                    ]
                },
                'url': {
                    '$not': re.compile(r"(?i)videos?")
                },
            }
        },
        {
            '$sort': {
                'createdAt': -1
            }
        },
        {
            '$group': {
                '_id': '$sourceLanguageName',
                'samples': {
                    '$push': {
                        'sourceId': '$sourceId',
                        'url': '$url',
                        'sourceLanguageName': '$sourceLanguageName'
                    }
                }
            }
        },
        {
            '$project': {
                'samples': {
                    '$slice': [
                        '$samples', 100
                    ]
                }
            }
        },
        {
            '$unwind': '$samples'
        },
        {
            '$replaceRoot': {
                'newRoot': '$samples'
            }
        },
        {
            '$sort': {
                'sourceId': -1
            }
        }
    ]

    data = MongoDB.aggregate(collection, pipeline)
    print(f"Fetched {len(data)} records.")

    data_update_collection = MongoDB.connect(
        mongo_uri,
        "ingestion-data",
        "summaries_test"
    )

    run_comparison(data, data_update_collection)
    return {"result": "success"}

print(f"Key:: {get_api_key()}")